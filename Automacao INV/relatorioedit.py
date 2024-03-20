from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

class Editor:

    def __init__(self, caminhoentrada, caminhosaida, nomearquivo) -> None:
        self.entrada = caminhoentrada
        self.saida = caminhosaida
        self.arquivo = nomearquivo

    def alterar_fonte(self, caminho, nomefonte, tamanho, saida, nomearquivo):
        caminho_planilha = load_workbook(caminho)
        planilha = caminho_planilha.active

        font = Font(name=f'{nomefonte}', size=tamanho)
        for row in planilha.iter_rows():
            for cell in row:
                cell.font = font
        caminho_planilha.save(f'{saida}\\{nomearquivo}')
        print(f'Arquivo salvo em: {saida}\\{nomearquivo}')


    def alterar_tamanho_coluna(planilha, coluna, tamanho):
        planilha.column_dimensions[f'{coluna}'].width = tamanho

    def colorir_linhas(linha, cor, caminho, arquivo, saida):

        caminho_planilha = load_workbook(caminho)
        fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        for cell in linha:
            cell.fill = fill

        caminho_planilha.save(f'{saida}\\{arquivo}')
        print(f'Arquivo salvo em: {saida}\\{arquivo}')
    
    
        



plan = Editor(r"C:\Users\rcramos2\Stefanini\teste\planilha_rog_colorida.xlsx", r"C:\Users\rcramos2\Stefanini\teste\teste2", "larissamanoela.xlsx")
#plan.alterar_fonte(caminho=plan.entrada, nomefonte='Arial', tamanho=37, saida=plan.saida, nomearquivo=plan.arquivo)

if 'TRANSITORIA' in str(row[0].value) and saldo_atual < 0.01:
    plan.colorir_linhas(str(row[0].value), "FF00FF", plan.entrada, plan.arquivo, plan.saida)

