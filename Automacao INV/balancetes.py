import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

caminho = rf"C:\Users\rcramos2\Stefanini\Célula 179 - Balancete INV\Balancetes\1 - Bronze\Balancetes_101083_CONTROLADORIA_OFICIAL_20240315.xls"

dados = pd.read_html(caminho)

cabeçalho = pd.DataFrame(dados[0])
balancete = pd.DataFrame(dados[1])
totais = pd.DataFrame(dados[2])

balancete_concat = balancete.dropna(subset=[0])
balancete_concat.columns = balancete_concat.iloc[0]
balancete_concat = balancete_concat[1:].reset_index(drop=True)

nome_coluna = [
    "CONTA",
    "SALDO ANTERIOR",
    "C/D1",
    "DEBITO",
    "CREDITO",
    "SALDO ATUAL",
    "C/D2",
]
balancete_concat.columns = nome_coluna

balancete_concat["SALDO ATUAL"] = balancete_concat["SALDO ATUAL"].str.replace(".", "").str.replace(",", ".").astype(float)

# Convertendo a coluna "SALDO ATUAL" para números
balancete_concat["SALDO ATUAL"] = pd.to_numeric(balancete_concat["SALDO ATUAL"], errors='ignore')

caminho_saida = rf"C:\Users\rcramos2\Stefanini\teste"
nome_arquivo_saida = 'balancete_concatenado_rogv.xlsx'


# Salvar o DataFrame como um arquivo Excel
balancete_concat.to_excel(f'{caminho_saida}\{nome_arquivo_saida}', index=False)

# Carregar o arquivo Excel
planilha_colorir = load_workbook(f'{caminho_saida}\{nome_arquivo_saida}')
planilha = planilha_colorir.active

# Definir a fonte para todas as células
font = Font(name='Calibri', size=9)
for row in planilha.iter_rows():
    for cell in row:
        cell.font = font

# Definir altura para todas as linhas
for row in planilha.iter_rows():
    planilha.row_dimensions[row[0].row].height = 15

planilha.column_dimensions['B'].width = 15
planilha.column_dimensions['D'].width = 15
planilha.column_dimensions['E'].width = 15
planilha.column_dimensions['F'].width = 15


green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for row in planilha.iter_rows(min_row=2, max_row=planilha.max_row, min_col=1, max_col=planilha.max_column):
    saldo_atual = row[5].value
    print("Valor da célula 'SALDO ATUAL':", saldo_atual)
    print("Valor da primeira célula da linha:", row[0].value)
    
    if 'TRANSITORIA' in str(row[0].value) and saldo_atual < 0.01:
        print("Condições atendidas para a linha (verde):", row)
        for cell in row:
            cell.fill = green_fill
    
    if 'TRANSITORIA' in str(row[0].value) and saldo_atual >= 0.01:
        print("Condições atendidas para a linha (vermelho):", row)
        for cell in row:
            cell.fill = red_fill


# Salvar as alterações no arquivo
planilha_colorir.save(f'{caminho_saida}\planilha_rog_colorida.xlsx')
